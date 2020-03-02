import os
import shutil
from SamsungWriter import samWriter
import openpyxl
import timeit

start = timeit.default_timer()



srcdir = '/Users/rupkumar.saha/Desktop/Ama_Files/'
destdir = '/Users/rupkumar.saha/Desktop/Sam_Files/'

def copyFolders(folder_name):
    if os.path.exists(destdir + folder_name):
        shutil.rmtree(destdir + folder_name)
    shutil.copytree(srcdir+folder_name, destdir+folder_name)
    modifyFolders()

def modifyFolders():
    for folders in os.listdir(destdir):
        if folders == '.DS_Store':
            continue
        wb = openpyxl.load_workbook(destdir+ folders + '/scraped_data.xlsx')
#        print(folders)
        sheet1 = wb.active
        for row in range(2, sheet1.max_row+1):
            cell = sheet1.cell(row=row, column=4)
            if cell.value is not None:
                cell.value = cell.value.replace('tag=glance09d-21',"")
        wb.save(destdir+ folders + '/scraped_data.xlsx')
        # for j in range(nrows-1):
        #     sheet.write(j+1, 3, sheet.cell_value(j+1, 3).replace('tag=glance09d-21',""))

for folders in os.listdir(srcdir):
    if folders == '.DS_Store':
        continue
#    print(folders)
    copyFolders(folders)
    samWriter()
    
stop = timeit.default_timer()

print('Time: ', stop - start)  